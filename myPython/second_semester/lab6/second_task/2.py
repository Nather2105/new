import openpyxl
from openpyxl.utils import get_column_letter

# Клас персон в який ми будемо закидувати ім'я та проекти кожної персони
class Person:
    def __init__(self, name, project):
        self.arr_of_projects = []
        self.arr_of_projects.append(project)
        self.name = name

# функція яка перевіряє чи є ім'я в масиві, якщо є то повертаємо його індекс
def is_include(arr, name):
    i = 0
    for person in arr:
        print("we comp it: " + person.name)
        print("with it: " + name)
        if(person.name == name):
            return i
        i += 1
    return None
    
# відкриваємо файл, переходимо на сторінку з інформацією та створюємо новий аркуш
wb = openpyxl.load_workbook("excel.xlsx")
ws = wb.create_sheet(title="improved")
wc = wb.active

# визначаємо стовпці з яких будемо брати ім'я та проекти
project_place = get_column_letter(1)
name_place = get_column_letter(2)
# масив із всіма персонами
arr_of_persons = []

for row in range(2, 1000):
    # перевірка на кінець інформації у файлі
    if(wc[name_place + str(row)].value == None):
        break
    # Витягуємо ім'я та проект персони із файлу
    name_of_person = wc[name_place + str(row)].value
    project_of_person = wc[project_place + str(row)].value
    # Перевіряємо чи є це ім'я у масиві 
    fl = is_include(arr_of_persons, name_of_person)
    # якщо немає то додаємо як нову персону
    # якщо є то додаємо новий проект в масив з проектами даної персони 
    if(fl != None):
        arr_of_persons[fl].arr_of_projects.append(project_of_person)
    else: 
        person = Person(name_of_person, project_of_person)
        arr_of_persons.append(person)

# переходимо на сторінку в якій будемо вводити оновлену таблицю 
info = ["Person1", "Person2", "Projects", "Weight"]
wimproved = wb["improved"]
# вставляємо тайтл
wimproved.append(info)
# збільшуємо колонку з проектами заради кращого виду
wimproved.column_dimensions['C'].width = 15
 
# проходимось по масиву з персонами знаходимо спільні проекти та вводимо у нову таблицю
for i in range(len(arr_of_persons)):
    for j in range(i+1, len(arr_of_persons)):
        person1 = arr_of_persons[i].name
        person2 = arr_of_persons[j].name
        projects = ''
        weight = 0
        fl = 1
        print("Person1 is " + person1)
        print("Person2 is " + person2)
        for proj in arr_of_persons[i].arr_of_projects:
            for proj1 in arr_of_persons[j].arr_of_projects:
                if(proj == proj1):
                    if(fl):
                        projects += proj
                        fl = 0
                    else:
                        projects += "_" + proj
                    weight += 1
                    
        print(projects)
        print(weight)
        wimproved.append([person1, person2, projects, weight])
                    
        print()
        
wb.save("updated.xlsx")

