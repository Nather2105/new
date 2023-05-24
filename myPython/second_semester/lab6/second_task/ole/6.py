import openpyxl
from openpyxl.utils import get_column_letter


class Provider:
    def __init__(self, id1):
        self.provider_id = id1
        self.provider_name = ''
        self.rating = 0
        self.address = ''
        
class Product:
    def __init__(self, id1):
        self.product_id = id1
        self.name = ''

class Order:
    def __init__(self, id1, id2):
        self.provider_id = id1
        self.product_id = id2
        self.price = 0
        self.term = 0
    
    

wb = openpyxl.load_workbook("lab6.xlsx")

arr_of_providers = []
arr_of_products = []
arr_of_orders = []

wprovider = wb["Постачальники"]
for row in range(2, 1000):
    
    
    col = get_column_letter(1)
    if(wprovider[col + str(row)].value == None):
        break
    
    id1 = wprovider[col + str(row)].value
    provider = Provider(id1)
    print("provider id is: " + provider.provider_id)
    
    col = get_column_letter(2)
    provider.provider_name = wprovider[col + str(row)].value
    print("provider name is: " + str(provider.provider_name))
    
    col = get_column_letter(3)
    provider.rating = wprovider[col + str(row)].value
    print("provider rating is: " + str(provider.rating))
    
    col = get_column_letter(4)
    provider.address = wprovider[col + str(row)].value
    print("provider id is: " + provider.address)
    
    print()
    arr_of_providers.append(provider)
    
print()

wprice = wb["Ціна"]
for row in range(2, 1000):
    
    col = get_column_letter(2)
    if(wprice[col + str(row)].value == None):
        break
    
    id1 = wprice[col + str(row)].value
    
    col = get_column_letter(3)
    id2 = wprice[col + str(row)].value
    order = Order(id1, id2)
    print("order provider id is: " + str(order.provider_id))
    print("order product id is: " + str(order.product_id))
    
    col = get_column_letter(4)
    price = wprice[col + str(row)].value
    order.price = price
    
    col = get_column_letter(5)
    term = wprice[col + str(row)].value
    order.term = term
    
    print("order price is: " + str(order.price))
    print("order term is: " + str(order.term))
    print()
    arr_of_orders.append(order)
    
print()
wproduct = wb["Продукція"]
for row in range(2, 1000):
    
    col = get_column_letter(1)
    if(wproduct[col + str(row)].value == None):
        break
    
    id1 = wproduct[col + str(row)].value
    product = Product(id1)
    print("product id is: " + str(product.product_id))
    
    col = get_column_letter(2)
    product.name = wproduct[col + str(row)].value
    print("product name is: " + str(product.name))
    print()
    arr_of_products.append(product)
    
    

a1 = 0.3
a2 = 0.7
p_max = 0
r_max = 0
tender = []
for provider in arr_of_providers:
    if(r_max < provider.rating):
        r_max = provider.rating
print("r_max = " + str(r_max))
for product in arr_of_products:
    tender_for_one_product = []
    for order in arr_of_orders:
        if(product.product_id == order.product_id):
            if(p_max < order.price):
                p_max = order.price
    print("p_max = " + str(p_max))
    
    
    for order in arr_of_orders:
        r = 0
        if(product.product_id == order.product_id):
            for provider in arr_of_providers:
                if(order.provider_id == provider.provider_id):
                    r = provider.rating
                    name = provider.provider_name
        if(r != 0):
            print("r = " + str(r))
            p = order.price
            print("p = " + str(p))
            s = float(a1) * float(p)/float(p_max) + float(a2) * float(r) / float(r_max)
            tender_for_one_product.append({'s' : s, 'provider' :  name, 'product_id' : order.product_id})
            print("s = " + str(s))
            print()
        
    tender.append(tender_for_one_product)
            
final_tender_for_products = []
for ten in tender:
    max_tender = 0
    for t in ten:
        if(t['s'] > max_tender):
            max_tender = t['s']
            name_of_prov = t['provider']
            id_of_prod = t['product_id']
            
    if(max_tender != 0):
        print("for product with id " + str(id_of_prod) + ", best tender is equal to " + str(max_tender) + ", provider of it is " + name_of_prov)
